"""
Excel导入导出工具类
"""
from typing import List, Dict, Any
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, date
import logging

logger = logging.getLogger(__name__)


class ExcelUtils:
    """Excel工具类"""
    
    @staticmethod
    def generate_grade_template(course_name: str = "示例课程") -> BytesIO:
        """
        生成成绩导入模板
        
        Args:
            course_name: 课程名称
            
        Returns:
            Excel文件的BytesIO对象
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "成绩导入模板"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30
        
        # 标题样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置表头
        headers = ["学号", "姓名", "分数", "备注"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 添加说明行
        ws.cell(row=2, column=1).value = "2024001"
        ws.cell(row=2, column=2).value = "张三"
        ws.cell(row=2, column=3).value = 85.5
        ws.cell(row=2, column=4).value = "进步明显"
        
        ws.cell(row=3, column=1).value = "2024002"
        ws.cell(row=3, column=2).value = "李四"
        ws.cell(row=3, column=3).value = 92.0
        ws.cell(row=3, column=4).value = ""
        
        # 添加填写说明
        ws.cell(row=5, column=1).value = "填写说明:"
        ws.cell(row=5, column=1).font = Font(bold=True, color="FF0000")
        
        ws.cell(row=6, column=1).value = "1. 学号: 必填,必须是已存在的学号"
        ws.cell(row=7, column=1).value = "2. 姓名: 必填,用于校验(与学号匹配)"
        ws.cell(row=8, column=1).value = "3. 分数: 必填,数字格式,范围0-100"
        ws.cell(row=9, column=1).value = "4. 备注: 选填,教师评语"
        ws.cell(row=10, column=1).value = f"5. 课程: {course_name}"
        ws.cell(row=11, column=1).value = "6. 考试类型、日期等信息在上传时填写"
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def parse_grade_excel(file_stream) -> List[Dict[str, Any]]:
        """
        解析成绩Excel文件
        
        Args:
            file_stream: Excel文件流
            
        Returns:
            成绩数据列表
            
        Raises:
            ValueError: Excel格式错误
        """
        try:
            # 将文件流读取到BytesIO对象中
            # 这样可以避免SpooledTemporaryFile的seekable问题
            file_content = file_stream.read()
            file_bytes = BytesIO(file_content)
            
            wb = load_workbook(file_bytes, data_only=True)
            ws = wb.active
            
            # 读取表头
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # 验证必需的列
            required_columns = ["学号", "分数"]
            for col in required_columns:
                if col not in headers:
                    raise ValueError(f"缺少必需的列: {col}")
            
            # 获取列索引
            student_code_idx = headers.index("学号")
            student_name_idx = headers.index("姓名") if "姓名" in headers else None
            score_idx = headers.index("分数")
            remark_idx = headers.index("备注") if "备注" in headers else None
            
            # 读取数据行
            grades_data = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # 跳过空行
                if not any(row):
                    continue
                
                # 跳过说明行(检查第一列是否包含"填写说明"或"说明")
                if row[0] and isinstance(row[0], str) and ("说明" in row[0] or "填写" in row[0]):
                    break
                
                student_code = row[student_code_idx] if student_code_idx < len(row) else None
                student_name = row[student_name_idx] if student_name_idx is not None and student_name_idx < len(row) else None
                score = row[score_idx] if score_idx < len(row) else None
                remark = row[remark_idx] if remark_idx is not None and remark_idx < len(row) else None
                
                # 跳过学号为空的行
                if not student_code:
                    continue
                
                grade_data = {
                    "student_code": str(student_code).strip(),
                    "student_name": str(student_name).strip() if student_name else None,
                    "score": score,
                    "remark": str(remark).strip() if remark else None
                }
                
                grades_data.append(grade_data)
            
            logger.info(f"Excel解析成功: 共{len(grades_data)}条数据")
            
            return grades_data
            
        except Exception as e:
            logger.error(f"Excel解析失败: {str(e)}")
            raise ValueError(f"Excel文件格式错误: {str(e)}")
    
    @staticmethod
    def export_grades_to_excel(grades: List[Dict[str, Any]], 
                              course_name: str = "课程") -> BytesIO:
        """
        导出成绩到Excel
        
        Args:
            grades: 成绩数据列表
            course_name: 课程名称
            
        Returns:
            Excel文件的BytesIO对象
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "成绩单"
        
        # 设置列宽
        column_widths = {
            'A': 12,  # 学号
            'B': 12,  # 姓名
            'C': 15,  # 考试类型
            'D': 20,  # 考试名称
            'E': 12,  # 分数
            'F': 12,  # 满分
            'G': 15,  # 考试日期
            'H': 30   # 备注
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 标题样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 添加标题行
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"{course_name} - 成绩单"
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        # 添加导出时间
        time_cell = ws.cell(row=2, column=1)
        time_cell.value = f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:H2')
        
        # 设置表头
        headers = ["学号", "姓名", "考试类型", "考试名称", "分数", "满分", "考试日期", "备注"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 填充数据
        for row_num, grade in enumerate(grades, start=4):
            ws.cell(row=row_num, column=1).value = grade.get("student_code", "")
            ws.cell(row=row_num, column=2).value = grade.get("student_name", "")
            ws.cell(row=row_num, column=3).value = grade.get("exam_type", "")
            ws.cell(row=row_num, column=4).value = grade.get("exam_name", "")
            ws.cell(row=row_num, column=5).value = grade.get("score", "")
            ws.cell(row=row_num, column=6).value = grade.get("full_score", 100)
            ws.cell(row=row_num, column=7).value = grade.get("exam_date", "")
            ws.cell(row=row_num, column=8).value = grade.get("remark", "")
            
            # 设置边框
            for col_num in range(1, 9):
                ws.cell(row=row_num, column=col_num).border = thin_border
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"成绩导出成功: 共{len(grades)}条数据")
        
        return output
